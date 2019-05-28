import os
import sys
import time
import ingest
import shutil
import openpyxl
import argparse
from datetime import datetime
import asnake.logging as logging
from asnake.client import ASnakeClient

parser = argparse.ArgumentParser()
parser.add_argument("ID", help="Collection ID for the files you are packaging.")
parser.add_argument("-p", "--path", help="Path of files to ingest.", default=None)
parser.add_argument("-a", "--accession", help="Optional ArchivesSpace Accession ID for new acquisisions.", default=None)
args = parser.parse_args()

print (str(datetime.now()) + " Transferring records for " + str(args.ID))

if args.path == None:
    path = os.path.join("/media/Library", args.ID.upper(), "transfer")
else:
    path = args.path
if not os.path.isdir(path):
    raise Exception("ERROR: " + str(path) + " is not a valid path.")

if not args.accession == None:
    print ("Using existing accession " + str(args.accession))
    accessionID = args.accession
else:
    # for automated accessions
    accessionProfilePath = "/media/Library/ESPYderivatives/accessionProfiles"
    accessionProfile = os.path.join(accessionProfilePath, args.ID.upper() + ".xlsx")
    if not os.path.isfile(accessionProfile):
        raise Exception("ERROR: Accession Profile not found at " + accessionProfile + ". Use an existing accession ID, or create a profile.")
    
    newAccession = {"external_ids":[], "related_accessions":[], "classifications":[], "subjects":[], "linked_events":[],\
    "extents":[], "dates":[], "external_documents":[], "rights_statements":[], "deaccessions":[], "related_resources":[],\
    "restrictions_apply":False, "access_restrictions":False, "use_restrictions":False, "linked_agents":[], "instances":[],\
    "id_0":"", "id_1":"", "title":"","content_description":"","condition_description":""}
    acquisitionTypes = ["deposit", "gift", "purchase", "transfer"]
    resourceTypes = ["collections", "papers", "publications", "records"]
    
    print ("Creating new accession using " + accessionProfile)
    wb = openpyxl.load_workbook(filename=accessionProfile, read_only=True)    
    for sheet in wb.worksheets:
        for row in sheet.rows:
            if not row[1].value == None:
                accessionKey = row[0].value.lower().replace(" ", "_")
                if accessionKey == "acquisition_type":
                    if not row[1].value.lower().strip() in acquisitionTypes:
                        raise Exception("ERROR: Failed to create automated accession. " + str(row[1].value) + " is not a valid acqusition type.")
                    else:
                        newAccession[accessionKey] = row[1].value.lower().strip()
                elif accessionKey == "resource_type":
                    if not row[1].value.lower().strip() in resourceTypes:
                        raise Exception("ERROR: Failed to create automated accession. " + str(row[1].value) + " is not a valid resource type.")
                    else:
                        newAccession[accessionKey] = row[1].value.lower().strip()
                else:
                    newAccession[accessionKey] = row[1].value
                if accessionKey == "access_restrictions_note":
                    newAccession["access_restrictions"] = True
                    newAccession["restrictions_apply"] = True
                elif accessionKey == "use_restrictions_note":
                    newAccession["use_restrictions"] = True
                    newAccession["restrictions_apply"] = True
    newAccession["accession_date"] = str(datetime.today().strftime('%Y-%m-%d'))
    year = newAccession["accession_date"].split("-")[0]
    
    #login to ASpace
    client = ASnakeClient()
    client.authorize()
    logging.setup_logging(stream=sys.stdout, level='INFO')
    
    # Get related resouces
    call = "repositories/2/search?type[]=resource&page=1&aq={\"query\":{\"field\":\"identifier\", \"value\":\"" + str(args.ID) + "\", \"jsonmodel_type\":\"field_query\"}}"
    resourceResponse = client.get(call).json()
    if len(resourceResponse["results"]) < 1:
        raise Exception("ERROR: Could not find resource with ID: " + str(args.ID))
    else:
        newAccession["related_resources"] = [{"ref": resourceResponse["results"][0]["uri"]}]
    
    # get accession id
    print ("Getting correct accession ID...")
    yearBegin = datetime.strptime(year + "-01-01 00:00", "%Y-%m-%d %H:%M")
    yearBeginPosix = str(time.mktime(yearBegin.timetuple())).split(".")[0]
    yearCall = "repositories/2/accessions?all_ids=true&modified_since=" + yearBeginPosix
    accessions = client.get(yearCall).json()
    
    idList = []
    for aID in accessions:
        entry = client.get("repositories/2/accessions/" + str(aID)).json()
        if entry["accession_date"].split("-")[0] == "2019":
            idList.append(int(entry["id_1"]))
    newID = max(idList) + 1
    if len(str(newID)) == 1:
        newID = "00" + str(newID)
    elif len(str(newID)) == 2:
        newID = "0" + str(newID)
    print ("Creating new accession " + year + "-" + str(newID) + "...")
    newAccession["id_0"] = year
    newAccession["id_1"] = str(newID)
    accessionID = year + "-" + str(newID)
    updateAccession = client.post("repositories/2/accessions", json=newAccession)
    if updateAccession.status_code == 200:
        print ("\tSuccessfully updated accession " + newAccession["id_0"] + "-" + newAccession["id_1"])
    else:
        print ("\tERROR " + str(updateAccession.status_code) + "! Failed to update accession: " + newAccession["id_0"] + "-" + newAccession["id_1"])
        
print ("Waiting for new accession to be indexed...")
time.sleep(120)
    
print ("Ingesting records at " + path)
arrangementSwitch = False
arrangementsPath = "/media/Masters/Archives/arrangements"
arrangements = os.path.join(arrangementsPath, args.ID.upper())
if os.path.isdir(arrangements):
    arrangementSwitch = True

SIP = ingest.main(args.ID, path, accessionID)

print ("Updating log files...")
newFiles = {}
fileInventory = SIP.inventory()
for newFile in fileInventory.split("\n"):
    if os.path.sep in newFile:
        root, filePath = newFile.split(os.path.sep, 1)
    else:
        root = "transferLog"
        filePath = newFile
    fullPath = os.path.join(SIP.data, newFile)
    modifiedTime = str(datetime.fromtimestamp(os.path.getmtime(fullPath)))
    if not root in newFiles.keys():
        newFiles[root] = []
    if arrangementSwitch == True:
        newFiles[root].append([os.path.dirname(filePath), os.path.basename(filePath), modifiedTime])
    else:
        newFiles[root].append([os.path.dirname(newFile), os.path.basename(filePath), modifiedTime])

logPath = os.path.join(os.path.dirname(path), "log")
if not os.path.isdir(logPath):
    os.mkdir(logPath)
for logGroup in newFiles.keys():
    excapedLogGroup = "".join([c for c in logGroup if c.isalpha() or c.isdigit() or c==' ']).rstrip()
    logFile = os.path.join(logPath, excapedLogGroup + ".xlsx")
    if os.path.isfile(logFile):
        logBook = openpyxl.load_workbook(filename=logFile, read_only=False)
        sheet = logBook.active
        startRow = int(sheet.max_row) + 1
    else:
        logBook = openpyxl.Workbook()
        sheet = logBook.active
        sheet["A1"] = "File Path"
        sheet["B1"] = "File name"
        sheet["C1"] = "Modified Date"
        startRow = 2
    for addedFile in newFiles[logGroup]:
        try:
            sheet["A" + str(startRow)] = addedFile[0]
            sheet["B" + str(startRow)] = addedFile[1]
            sheet["C" + str(startRow)] = addedFile[2]
            startRow += 1
        except:
            logBook.save(filename=logFile)
        
    logBook.save(filename=logFile)
    print ("\tUpdated " + os.path.basename(logFile))


# set up transfer directory again
os.mkdir(path)
if os.path.isdir(arrangements):
    for folder in os.listdir(arrangements):
        shutil.copytree(os.path.join(arrangements, folder), os.path.join(path, folder))
